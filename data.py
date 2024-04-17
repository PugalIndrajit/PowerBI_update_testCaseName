# data.py(working update testcase name script)
# ==============================================


import openpyxl  # For Excel Generation
import sys
import re

# import time
import requests
import json
import numpy as np  # For Percentile 75
import argparse
import os

STB_TOKEN = "sTMoi6j9Z2V6wmOabhIXgylMoq2m59b_"
api_url = "http://innowave.stb-tester.com/api/v2/results"


class Data:
    def __init__(self):
        self.api_url = "http://innowave.stb-tester.com/api/v2"

    # ------------------------------------------------------------- Method to extract measurement list from response------------------------------------------------------------------------------------------------
    def measure_list(self, response, type):
        """
        extracting the measurement list from the response based on the type
        """

        lines = response.text.split("\n")
        m_list = []
        null_value = [[0.0]]

        for line in lines:
            if type in line:
                pattern1 = r"\s*\[([^\]]+)\]"
                pattern = rf"{type}{pattern1}"
                match = re.search(pattern, line)
                if match:
                    measurements_str = match.group(1)
                    current_measurements = [
                        float(value.strip()) for value in measurements_str.split(",")
                    ]
                    m_list.append(current_measurements)

        if len(m_list) > 0:
            return m_list
        else:
            return null_value

    # ---------------------------------------------------- Method to fetch static data from response-----------------------------------------------------#
    def measure_static_data(self, response, type):
        """
        getting the static data from the response by splitting the response into the different lines
        and findou the pattern as the next line based on the type and return
        """

        lines = response.text.split("\n")
        out = ""
        for line in lines:
            if type in line:
                pattern1 = r"([^\n]+)"
                pattern = rf"{type}{pattern1}"
                match = re.search(pattern, line)
                if match:
                    measurements_str = match.group(1).split()[0]
                    return measurements_str
        return out

    # ----------------------------------------------Fetching Log Output (Measurements)------------------------------#
    def fetch_log_output(self, result_id):
        """
        by taking the result id and the toke we call the apis which will give the stbt.log  along with that
        we called the measure list method to extract measure list, cpu list, ram_list and data dict by adding
        all of them into a single list return it as a data
        """

        headers = {"Authorization": "token " + STB_TOKEN}
        params = {"tz": "Asia/Calcutta"}
        nullValue = [[0.0]]
        m_list = []
        cpu_list = []
        ram_list = []
        data_dict = {}
        stb_fw = ""
        stb_version = ""
        android_version = ""
        try:
            response = requests.get(
                self.api_url + "/results" + result_id + "/stbt.log",
                headers=headers,
                params=params,
            )
            if response.status_code != 200:
                print(response.text)
                sys.exit(1)
            else:
                m_list = self.measure_list(response, "measurements: ")
                cpu_list = self.measure_list(response, "CPU info : TELEKOM PACKAGE =  ")
                ram_list = self.measure_list(response, "RAM info : TOTAL = ")
                stb_fw = self.measure_static_data(response, "Build Version: ")
                stb_version = self.measure_static_data(
                    response, "Version Number of One TV     versionName="
                )
                android_version = self.measure_static_data(
                    response, "Android OS Version: "
                )
            if len(m_list) > 0 or len(cpu_list) > 0 or len(ram_list) > 0:
                data_dict["cpu_list"] = cpu_list
                data_dict["ram_list"] = ram_list
                data_dict["m_list"] = m_list
                data_dict["stb_fw"] = stb_fw
                data_dict["stb_version"] = stb_version
                data_dict["android_version"] = android_version
                return data_dict
            else:
                return nullValue

        except Exception as e:
            print(e)
            sys.exit(1)

    # ----------------------------------------------Method to fetch data from API based on job ID--------------------------------------------------------------------#
    def fetch_data_from_api(self, job_id):
        """
        method used to fetch the data from the apis or raise the error
        """
        headers = {"Authorization": "token " + STB_TOKEN}
        params = {"filter": f"job:{job_id}"}
        try:
            response = requests.get(api_url, headers=headers, params=params)
            response.raise_for_status()  # Raise an exception for bad status codes
            return response.json()
        except requests.exceptions.RequestException as e:
            print("Error fetching data from API:", e)
            return None

    # ----------------------------------------------- Method to curate dataset from log data------------------------------------------------------------------------------------------#
    def dataset_curation(self, log_data):
        """
        this method probably takes the data dict from fetch log output  and  does loop the result and filter all the
        job id results of the cpu, ram, measur... into seperate lists from different each response and aggregate them into
        single lists  and store them into the respective nested lists add them into a new dictionary and returning it
        """
        data_dict = {}
        measurement_list = []
        cpu_list = []
        ram_list = []
        stb_fw_list = []
        stb_version_list = []
        android_version_list = []

        for result_id, data_dictionary in log_data.items():
            # data_dictionary=self.fetch_log_output(result_id)
            measurements = data_dictionary["m_list"]
            measurement_list.append(measurements)
            cpu = data_dictionary["cpu_list"]
            cpu_list.append(cpu)
            ram = data_dictionary["ram_list"]
            ram_list.append(ram)
            stb_fw = data_dictionary["stb_fw"]
            stb_fw_list.append(stb_fw)
            stb_version = data_dictionary["stb_version"]
            stb_version_list.append(stb_version)
            android_version = data_dictionary["android_version"]
            android_version_list.append(android_version)

        data_dict["cpu_list"] = cpu_list
        data_dict["ram_list"] = ram_list
        data_dict["m_list"] = measurement_list
        data_dict["stb_fw"] = stb_fw_list
        data_dict["stb_version"] = stb_version_list
        data_dict["android_version"] = android_version_list
        return data_dict

    # --------------------------------------------------------- Method to extract dataset from API data and curated data-----------------------------------------------------------#
    def dataset_extraction(self, data, data_dict, node_id):
        # headers = {"Authorization": "token " + STB_TOKEN}
        headers = {"Authorization": "token mFSw7tQLp5Yu43IlDwU99crL3EiEZl4T"}

        response = requests.get(self.api_url + "/nodes/" + node_id, headers=headers)
        if response.status_code != 200:
            print(response.text)
            sys.exit(1)

        parsed_data = json.loads(response.text)
        print("Natco Name:" + parsed_data["config"]["node"]["friendly_name"])
        file_path_out = f"STB_Database_{node_id}.xlsx"
        # Create a new workbook and select the active worksheet

        # wb = openpyxl.load_workbook('STB_Database.xlsx')
        wb = openpyxl.Workbook()  # Create New WB
        ws = wb.active

        # Write headers to the worksheet
        headers = [
            "Run Type",
            "Date",
            "Iteration Number",
            "Test Case",
            "Load Time",
            "CPU",
            "RAM",
            "start_time",
            "end_time",
            "Job UID",
            "Node Id",
            "Failure Reason",
            "Result",
            "Natco",
            "Percentile 75 Load Times",
            "Percentile 75 CPU",
            "Percentile 75 RAM",
            "Country Code",
            "STB Release",
            "STB Firmware",
            "STB Android Version",
            "STB Build Info",
            "NatcoNode",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        startingIndex = (
            ws.max_row + 1
        )  # index to determine from where to push data in powerbi
        first = True
        second = False
        third = False
        fourth = False
        fifth = False
        load_times1 = []
        cpu_times1 = []
        ram_times1 = []
        in1 = 0
        measurements = data_dict["m_list"]
        cpu = data_dict["cpu_list"]
        ram = data_dict["ram_list"]
        stb_version = data_dict["stb_version"][-1]
        android_version = data_dict["android_version"][-1]
        if not android_version:
            android_version = "11"
        stb_fw = data_dict["stb_fw"][-1]
        if not stb_version:
            stb_version = "2.0.620"
        in1Max = len(cpu)

        runType = (
            "rerun"  # Type of Run   AT MR2.5,2.6,649  AT MR2.5, MR2.6 and 649 - rerun
        )

        # Populate the worksheet with data and load times
        for i, dat in enumerate(data, start=1):
            # print("in1:",in1)
            if in1 < in1Max:
                cpu_times1 = cpu[in1]
                ram_times1 = ram[in1]
            # print("Cpu Times 1:",cpu_times1)
            # print("Ram Times 1:",ram_times1)

            test_case = dat["test_case"].split("::")[1]
            startTime = dat["start_time"].split("T")[1]
            endTime = dat["end_time"].split("T")[1]
            job_uid = dat["job_uid"]
            nodeId = dat["job_uid"].split("/")[1]
            failureReason = dat["failure_reason"]
            Result = dat["result"]
            date = dat["start_time"].split("T")[0]
            natcoName = parsed_data["config"]["node"]["friendly_name"]
            countryCode = parsed_data["config"]["device_under_test"]["language"]
            natcoNode = countryCode.upper()
            buildInfo = f"{natcoName} {stb_version} {stb_fw} {android_version}"
            if countryCode == "MKT":
                countryCode = "MKD"
            stbRelease = stb_version

            load_times1 = measurements[i - 1]  # Load times for the current test case
            # print("measurement i-1:",load_times1 ," i:", i)
            index = 0
            indexMax = len(cpu_times1)

            # print("Level-1:",load_times1)
            for k, load_row in enumerate(load_times1):
                # print("index:",index, " indexMax:",indexMax)
                load_times = load_row
                if index < indexMax:
                    cpu_times = cpu_times1[index]
                    ram_times = ram_times1[index]
                    # print("Cpu Times 2=",cpu_times)
                    # print("Ram Times 2=",ram_times)

                if (
                    "performance08_ott_zapping.py::test_ott_zapping_channel_up_and_down"
                    in dat["test_case"]
                ):
                    test_case_up = "OTT Zapping Channel UP"
                    test_case_down = "OTT Zapping Channel Down"
                elif (
                    "performance19_epg_navigation.py::test_epg_navigation"
                    in dat["test_case"]
                ):
                    test_case_up = "EPG Zapping Channel UP"
                    test_case_down = "EPG Zapping Channel Down"
                    test_case_left = "EPG Zapping Left"
                    test_case_right = "EPG Zapping Right"
                elif (
                    "performance19_epg_navigation_HU.py::test_epg_navigation_HU"
                    in dat["test_case"]
                ):
                    test_case_up = "EPG Zapping HU Channel UP"
                    test_case_down = "EPG Zapping HU Channel Down"
                    test_case_left = "EPG Zapping HU Left"
                    test_case_right = "EPG Zapping HU Right"

                # Processing OTT Zapping Channel UP and Down
                if (
                    "performance08_ott_zapping.py::test_ott_zapping_channel_up_and_down"
                    in dat["test_case"]
                ):
                    test_case = test_case_up if first else test_case_down
                    first = not first

                # Processing EPG Zapping
                elif (
                    "performance19_epg_navigation.py::test_epg_navigation"
                    in dat["test_case"]
                    or "performance19_epg_navigation_HU.py::test_epg_navigation_HU"
                    in dat["test_case"]
                ):
                    if second == 0:
                        test_case = test_case_up
                    elif second == 1:
                        test_case = test_case_down
                    elif second == 2:
                        test_case = test_case_left
                    elif second == 3:
                        test_case = test_case_right
                    second = (second + 1) % 4
                #    # Initialize variables for test cases
                #     test_case_up = test_case_down = test_case_left = test_case_right = None

                #     # Define test cases mapping
                #     test_cases_mapping = {
                #        'performance08_ott_zapping.py::test_ott_zapping_channel_up_and_down': ('OTT Zapping Channel UP', 'OTT Zapping Channel Down'),
                #        'performance19_epg_navigation.py::test_epg_navigation': ('EPG Zapping Channel UP', 'EPG Zapping Channel Down', 'EPG Zapping Left', 'EPG Zapping Right'),
                #         'performance19_epg_navigation_HU.py::test_epg_navigation_HU': ('EPG Zapping HU Channel UP', 'EPG Zapping HU Channel Down', 'EPG Zapping HU Left', 'EPG Zapping HU Right')
                #     }

                #     # Iterate through the test cases mapping to find the matching test case
                #     for test_case_name, values in test_cases_mapping.items():
                #         if test_case_name in dat['test_case']:
                #             if 'HU' in test_case_name:
                #                 test_case_up, test_case_down, test_case_left, test_case_right = values
                #             else:
                #                 test_case_up, test_case_down = values[:2]  # Ensure only two values are unpacked
                #             break
                #      # Processing OTT Zapping Channel UP and Down
                #     if 'performance08_ott_zapping.py::test_ott_zapping_channel_up_and_down' in dat['test_case']:
                #         test_case = test_case_up if first else test_case_down
                #         first = not first

                #     # Processing EPG Zapping
                #     elif 'performance19_epg_navigation.py::test_epg_navigation' in dat['test_case'] or 'performance19_epg_navigation_HU.py::test_epg_navigation_HU' in dat['test_case']:
                #         if second == 0:
                #             test_case = test_case_up
                #         elif second == 1:
                #             test_case = test_case_down
                #         elif second == 2:
                #             test_case = test_case_left
                #         elif second == 3:
                #             test_case = test_case_right
                #         second = (second + 1) % 4

                percentileLoad75 = np.percentile(load_times, 75)
                percentileCPU75 = np.percentile(cpu_times, 75)
                percentileRAM75 = np.percentile(ram_times, 75)
                iteration = 1
                # print("Level-2:",load_times)
                index2 = 0
                index2Max = len(cpu_times)
                testcase_dict = {
                    "test_main_menu_page_load_performance": "MAIN MENU PERFORMANCE",
                    "test_epg_home_menu_load_performance": "FULL EPG HOME MENU PERFORMANCE",
                    "test_epg_horizontal_right_pagination_performance": "EPG HORIZONTAL RIGHT PAGINATION",
                    "test_epg_horizontal_left_pagination_performance": "EPG HORIZONTAL LEFT PAGINATION",
                    "test_epg_vertical_down_pagination_performance": "EPG VERTICAL DOWN PAGINATION",
                    "test_epg_vertical_up_pagination_performance": "EPG VERTICAL UP PAGINATION",
                    "test_epg_navigation": "EPG NAVIGATION",
                    "test_epg_live_load_performance": "FULL EPG LIVE PERFORMANCE",
                    "test_livetv_load_performance": "LIVE TV PERFORMANCE",
                    "OTT Zapping Channel UP": "OTT ZAPPING(CHANNEL DOWN) PERFORMANCE",
                    "OTT Zapping Channel Down": "OTT ZAPPING(CHANNEL UP) PERFORMANCE",
                    "test_ott_zapping_numpad_performance": "OTT ZAPPING(NUMPAD) PERFORMANCE",
                    "test_channel_up_for_25times": "CHANNEL UP(25 TIMES) PERFORMANCE",
                    "test_channel_down_for_25times": "CHANNEL DOWN(25 TIMES) PERFORMANCE",
                    "test_zapp_bar_load_performance": "ZAPP BAR PERFORMANCE",
                    "test_standby_live_tv_load_performance": "STANDBY LIVE TV PERFORMANCE",
                    "test_search_result_page_performance": "SEARCH CONTENT PERFORMANCE",
                    "test_vod_details_performance": "VOD DETAILS PAGE PERFORMANCE",
                    "test_vod_episodes_details_performance": "VOD EPISODE DETAILS PAGE PERFORMANCE",
                    "test_vod_playback_performance": "VOD PLAYBACK PERFORMANCE",
                    "test_get_serial_no": "TEST SERIAL NO",
                    "test_get_client_id": "TEST CLIENT ID",
                    "test_apk_and_build_versions_info": "APK & BUILD VERSIONS INFO",
                    "EPG Zapping LEFT": "EPG NAVIGATION CHANNEL LEFT",
                    "EPG Zapping RIGHT": "EPG NAVIGATION CHANNEL RIGHT",
                    "EPG Zapping Channel UP": "EPG NAVIGATION CHANNEL UP",
                    "EPG Zapping Channel DOWN": "EPG NAVIGATION CHANNEL DOWN",
                    "EPG Zapping HU Left": "EPG NAVIGATION CHANNEL LEFT",
                    "EPG Zapping HU Right": "EPG NAVIGATION CHANNEL RIGHT",
                    "EPG Zapping HU Channel UP": "EPG NAVIGATION CHANNEL UP",
                    "EPG Zapping HU Channel Down": "EPG NAVIGATION CHANNEL DOWN",
                    "test_apk_and_build_versions_info_HU_only": "HU APK & BUILD VERSIONS INFO",
                }
                if test_case in testcase_dict:
                    test_case = testcase_dict[test_case]
                # Iterate through load times for the current test case
                for j, load_time in enumerate(load_times):
                    # print("index2:",index2)
                    cpu_value = 0.0
                    ram_value = 0.0
                    if index2 < index2Max:
                        cpu_value = cpu_times[index2]
                        ram_value = ram_times[index2]
                        # print("CPU Value inside last for loop:",cpu_value)
                        # print("Ram value inside last for loop",ram_value)
                    next_row = len(ws["A"]) + 1  # Get the next row number
                    # Populate the row with data
                    ws.cell(row=next_row, column=1, value=runType)
                    ws.cell(row=next_row, column=2, value=date)
                    ws.cell(row=next_row, column=3, value=iteration)  # Iteration Number
                    ws.cell(row=next_row, column=4, value=test_case)  # Test Case
                    ws.cell(row=next_row, column=5, value=load_time)  # Load Time
                    ws.cell(row=next_row, column=6, value=cpu_value)  # CPU Time
                    ws.cell(row=next_row, column=7, value=ram_value)  # RAM Time
                    ws.cell(
                        row=next_row, column=8, value=startTime.split("+")[0]
                    )  # Build
                    ws.cell(
                        row=next_row, column=9, value=endTime.split(".")[0]
                    )  # End Time
                    ws.cell(row=next_row, column=10, value=job_uid)  # Job UID
                    ws.cell(row=next_row, column=11, value=nodeId)  # Node Id
                    ws.cell(
                        row=next_row, column=12, value=failureReason
                    )  # Failure Reason
                    ws.cell(row=next_row, column=13, value=Result)  # Result
                    ws.cell(row=next_row, column=14, value=natcoName)  # Natco Name
                    ws.cell(
                        row=next_row, column=15, value=percentileLoad75
                    )  # Percentile75 Load
                    ws.cell(
                        row=next_row, column=16, value=percentileCPU75
                    )  # Percentile75  CPU
                    ws.cell(
                        row=next_row, column=17, value=percentileRAM75
                    )  # Percentile75  RAM
                    ws.cell(row=next_row, column=18, value=countryCode)  # Country Code
                    ws.cell(row=next_row, column=19, value=stbRelease)  # STB Version
                    ws.cell(row=next_row, column=20, value=stb_fw)  # STB FW
                    ws.cell(
                        row=next_row, column=21, value=android_version
                    )  # STB Android
                    ws.cell(row=next_row, column=22, value=buildInfo)  # STB Build
                    ws.cell(row=next_row, column=23, value=natcoNode)  # STB Natco Node
                    iteration = iteration + 1
                    index2 += 1
                index += 1
            in1 += 1
        # Save the workbook to a file
        endingIndex = ws.max_row

        wb.save(file_path_out)  # Main DB

        # JSON for PowerBI
        data_list = []
        for row_num in range(startingIndex, endingIndex):
            row_data = {
                "Run Type": ws.cell(row=row_num, column=1).value,
                "Date": ws.cell(row=row_num, column=2).value,
                "Iteration Number": ws.cell(row=row_num, column=3).value,
                "Test Case": ws.cell(row=row_num, column=4).value,
                "Load Time": ws.cell(row=row_num, column=5).value,
                "CPU": ws.cell(row=row_num, column=6).value,
                "RAM": ws.cell(row=row_num, column=7).value,
                "start_time": ws.cell(row=row_num, column=8).value,
                "end_time": ws.cell(row=row_num, column=9).value,
                "Job UID": ws.cell(row=row_num, column=10).value,
                "Node Id": ws.cell(row=row_num, column=11).value,
                "Failure Reason": ws.cell(row=row_num, column=12).value,
                "Result": ws.cell(row=row_num, column=13).value,
                "Natco": ws.cell(row=row_num, column=14).value,
                "Percentile 75 Load Times": ws.cell(row=row_num, column=15).value,
                "Percentile 75 CPU": ws.cell(row=row_num, column=16).value,
                "Percentile 75 RAM": ws.cell(row=row_num, column=17).value,
                "Country Code": ws.cell(row=row_num, column=18).value,
                "STB Release": ws.cell(row=row_num, column=19).value,
                "STB Firmware": ws.cell(row=row_num, column=20).value,
                "STB Android Version": ws.cell(row=row_num, column=21).value,
                "STB Build Info": ws.cell(row=row_num, column=22).value,
                "NatcoNode": ws.cell(row=row_num, column=23).value,
            }
            data_list.append(row_data)

        return data_list


api_instance = Data()
# job_id = "/stb-tester-48b02d5b0ab2/0000/1084" # Example job ID
# job_id = "/stb-tester-48b02da84b3e/0000/1808"
# job_id = "/stb-tester-48b02da8446d/0000/1742"
job_id = "/stb-tester-48b02da8446d/0000/1741"
data = api_instance.fetch_data_from_api(job_id)

# Check if API data is fetched successfully
if data:
    # Initialize an empty dictionary to store log data for each result ID
    log_data = {}
    # Iterate over the result IDs and call fetch_log_output for each result ID
    for result in data:
        # Extract the result ID
        result_id = result["result_id"]

        # Call the fetch_log_output method with the result ID
        log_data[result_id] = api_instance.fetch_log_output(result_id)
        # print(log_data)

    curated_data = api_instance.dataset_curation(log_data)
    # print(curated_data)
    # Define the node ID
    node_id = "stb-tester-48b02da84b3e"  # Example node ID

    # Call the dataset_extraction method with the required parameters
    extracted_data = api_instance.dataset_extraction(data, curated_data, node_id)

    # Print or further process the extracted data
    # print(extracted_data)
